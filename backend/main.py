from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import json
import io
import os
import boto3
from botocore.config import Config

from database import engine, get_db, Base
from models import Device
from extractor import extract_specs
from storage import upload_pdf, delete_pdf

Base.metadata.create_all(bind=engine)

app = FastAPI(title="RF Datasheet Tool")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=False,
)

DEVICE_TYPES = ["PA", "LNA", "Filter", "Switch", "FEM", "Balun", "Splitter", "RF-Connector"]


def parse_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(str(value).replace("~", "").strip())
    except (ValueError, TypeError):
        return None


def get_cos_client():
    region = os.getenv("COS_REGION", "ap-guangzhou")
    return boto3.client(
        "s3",
        endpoint_url=f"https://cos.{region}.myqcloud.com",
        aws_access_key_id=os.getenv("COS_SECRET_ID"),
        aws_secret_access_key=os.getenv("COS_SECRET_KEY"),
        config=Config(signature_version="s3v4", s3={"addressing_style": "virtual"}),
        region_name=region,
    )


@app.post("/api/upload")
async def upload_datasheet(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are supported")

    pdf_bytes = await file.read()

    try:
        stored_name, pdf_url = upload_pdf(pdf_bytes, file.filename)
    except Exception as e:
        raise HTTPException(500, f"Upload failed: {str(e)}")

    try:
        result = extract_specs(pdf_bytes)
    except Exception as e:
        try:
            delete_pdf(stored_name)
        except Exception:
            pass
        raise HTTPException(500, f"Extraction failed: {str(e)}")

    bands = result.get("bands") or []
    freq_min = result.get("freq_min_mhz")
    freq_max = result.get("freq_max_mhz")
    if not freq_min and bands:
        freq_min = min((b.get("freq_min_mhz") or 0) for b in bands if b.get("freq_min_mhz"))
    if not freq_max and bands:
        freq_max = max((b.get("freq_max_mhz") or 0) for b in bands if b.get("freq_max_mhz"))

    device = Device(
        name=result.get("device_name", file.filename),
        manufacturer=result.get("manufacturer"),
        device_type=result.get("device_type"),
        freq_min_mhz=freq_min,
        freq_max_mhz=freq_max,
        package=result.get("package"),
        package_size=result.get("package_size"),
        pin_count=result.get("pin_count"),
        enable_level=result.get("enable_level"),
        switch_logic=result.get("switch_logic"),
        bands=bands,
        pdf_filename=stored_name,
        pdf_url=pdf_url,
        raw_specs=result,
    )
    db.add(device)
    db.commit()
    db.refresh(device)

    return {"id": device.id, "name": device.name, "device_type": device.device_type}


@app.get("/api/devices")
def list_devices(
    device_type: Optional[str] = Query(None),
    freq_min: Optional[float] = Query(None),
    freq_max: Optional[float] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Device)
    if device_type:
        q = q.filter(Device.device_type == device_type)
    if freq_min is not None:
        q = q.filter(Device.freq_max_mhz >= freq_min)
    if freq_max is not None:
        q = q.filter(Device.freq_min_mhz <= freq_max)
    devices = q.order_by(Device.device_type, Device.name).all()

    return [_device_to_dict(d) for d in devices]


@app.get("/api/devices/{device_id}")
def get_device(device_id: int, db: Session = Depends(get_db)):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(404, "Device not found")
    return _device_to_dict(device)


@app.delete("/api/devices/{device_id}")
def delete_device(device_id: int, db: Session = Depends(get_db)):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(404, "Device not found")
    if device.pdf_filename:
        try:
            delete_pdf(device.pdf_filename)
        except Exception:
            pass
    db.delete(device)
    db.commit()
    return {"ok": True}


@app.get("/api/devices/{device_id}/datasheet")
def download_datasheet(device_id: int, db: Session = Depends(get_db)):
    """Proxy download of the original PDF from COS (forced download)."""
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(404, "Device not found")
    if not device.pdf_filename:
        raise HTTPException(404, "No PDF on file")

    bucket = os.getenv("COS_BUCKET_NAME")
    cos = get_cos_client()
    try:
        obj = cos.get_object(Bucket=bucket, Key=device.pdf_filename)
        pdf_bytes = obj["Body"].read()
    except Exception as e:
        raise HTTPException(500, f"Failed to fetch PDF: {str(e)}")

    filename = device.name + ".pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/devices/{device_id}/preview")
def preview_datasheet(device_id: int, db: Session = Depends(get_db)):
    """Proxy PDF inline for browser preview."""
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(404, "Device not found")
    if not device.pdf_filename:
        raise HTTPException(404, "No PDF on file")

    bucket = os.getenv("COS_BUCKET_NAME")
    cos = get_cos_client()
    try:
        obj = cos.get_object(Bucket=bucket, Key=device.pdf_filename)
        pdf_bytes = obj["Body"].read()
    except Exception as e:
        raise HTTPException(500, f"Failed to fetch PDF: {str(e)}")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{device.name}.pdf"'},
    )


@app.get("/api/export")
def export_devices(
    device_type: Optional[str] = Query(None),
    freq_min: Optional[float] = Query(None),
    freq_max: Optional[float] = Query(None),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    q = db.query(Device)
    if device_type:
        q = q.filter(Device.device_type == device_type)
    if freq_min is not None:
        q = q.filter(Device.freq_max_mhz >= freq_min)
    if freq_max is not None:
        q = q.filter(Device.freq_min_mhz <= freq_max)
    devices = q.order_by(Device.device_type, Device.name).all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    type_headers = {
        "PA": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名", "Vcc(V)", "Icc(mA)",
               "增益Typ(dB)", "增益Min(dB)", "P1dB(dBm)", "Psat(dBm)", "PAE(%)",
               "S11(dB)", "S22(dB)", "使能电平", "备注"],
        "LNA": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名", "Vcc(V)", "Icc(mA)",
                "增益Typ(dB)", "增益Min(dB)", "NF Typ(dB)", "NF Max(dB)",
                "IIP3(dBm)", "OP1dB(dBm)", "OIP3(dBm)", "S11(dB)",
                "使能电平", "开关逻辑", "备注"],
        "Filter": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名",
                   "IL Typ(dB)", "IL Max(dB)", "带外抑制(dB)", "回波损耗(dB)", "备注"],
        "Switch": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名", "Vcc(V)",
                   "IL Typ(dB)", "IL Max(dB)", "隔离Typ(dB)", "隔离Min(dB)",
                   "P1dB(dBm)", "功率容量(dBm)", "端口", "使能电平", "开关逻辑", "备注"],
        "FEM": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名", "Vcc(V)", "Icc(mA)",
                "TX增益(dB)", "TX P1dB(dBm)", "TX Psat(dBm)",
                "RX增益(dB)", "RX NF(dB)", "端口", "使能电平", "开关逻辑", "备注"],
        "Balun": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名",
                  "IL Typ(dB)", "回波损耗(dB)", "幅度不平衡(dB)", "相位不平衡(°)", "阻抗(Ω)", "备注"],
        "Splitter": ["型号", "厂家", "封装", "封装尺寸", "Pin数", "频段(MHz)", "频段名",
                     "IL Typ(dB)", "回波损耗(dB)", "隔离(dB)", "幅度不平衡(dB)", "相位不平衡(°)",
                     "功率容量(dBm)", "端口", "备注"],
        "RF-Connector": ["型号", "厂家", "封装/型号", "封装尺寸", "频段(MHz)", "频段名",
                         "IL Typ(dB)", "回波损耗(dB)", "VSWR", "阻抗(Ω)", "功率容量(dBm)", "备注"],
    }

    for dtype in DEVICE_TYPES:
        devs = [d for d in devices if d.device_type == dtype]
        if not devs:
            continue
        ws = wb.create_sheet(dtype)
        headers = type_headers.get(dtype, ["型号", "厂家", "封装", "Pin数", "频段(MHz)", "频段名", "备注"])
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(len(h) * 1.8, 12)

        row = 2
        for d in devs:
            bands = d.bands or [{}]
            switch_str = ""
            if d.switch_logic:
                switch_str = "; ".join(
                    f"{s.get('ctrl','')}→{s.get('path','')}" for s in d.switch_logic
                )
            notes = d.raw_specs.get("notes", "") if d.raw_specs else ""

            for band in bands:
                freq_str = f"{band.get('freq_min_mhz','')}–{band.get('freq_max_mhz','')}"
                bn = band.get("band_name") or ""

                if dtype == "PA":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("vcc_v"), band.get("icc_ma"),
                                band.get("gain_db"), band.get("gain_min_db"),
                                band.get("p1db_dbm"), band.get("psat_dbm"), band.get("pae_percent"),
                                band.get("s11_db"), band.get("s22_db"), d.enable_level, notes]
                elif dtype == "LNA":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("vcc_v"), band.get("icc_ma"),
                                band.get("gain_db"), band.get("gain_min_db"),
                                band.get("nf_db"), band.get("nf_max_db"),
                                band.get("iip3_dbm"), band.get("op1db_dbm"), band.get("oip3_dbm"),
                                band.get("s11_db"), d.enable_level, switch_str, notes]
                elif dtype == "Filter":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("insertion_loss_db"), band.get("insertion_loss_max_db"),
                                band.get("rejection_db"), band.get("return_loss_db"), notes]
                elif dtype == "Switch":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("vcc_v"),
                                band.get("insertion_loss_db"), band.get("insertion_loss_max_db"),
                                band.get("isolation_db"), band.get("isolation_min_db"),
                                band.get("p1db_dbm"), band.get("power_handling_dbm"), band.get("ports"),
                                d.enable_level, switch_str, notes]
                elif dtype == "FEM":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("vcc_v"), band.get("icc_ma"),
                                band.get("tx_gain_db"), band.get("tx_p1db_dbm"), band.get("tx_psat_dbm"),
                                band.get("rx_gain_db"), band.get("rx_nf_db"),
                                band.get("ports"), d.enable_level, switch_str, notes]
                elif dtype == "Balun":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("insertion_loss_db"), band.get("return_loss_db"),
                                band.get("amplitude_balance_db"), band.get("phase_balance_deg"),
                                band.get("impedance_ohm"), notes]
                elif dtype == "Splitter":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn,
                                band.get("insertion_loss_db"), band.get("return_loss_db"),
                                band.get("isolation_db"),
                                band.get("amplitude_balance_db"), band.get("phase_balance_deg"),
                                band.get("power_handling_dbm"), band.get("ports"), notes]
                elif dtype == "RF-Connector":
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, freq_str, bn,
                                band.get("insertion_loss_db"), band.get("return_loss_db"),
                                band.get("vswr"), band.get("impedance_ohm"),
                                band.get("power_handling_dbm"), notes]
                else:
                    row_data = [d.name, d.manufacturer, d.package, d.package_size, d.pin_count, freq_str, bn, notes]

                for col, val in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=val)
                row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rf_devices.xlsx"},
    )


def _device_to_dict(d: Device) -> dict:
    return {
        "id": d.id,
        "name": d.name,
        "manufacturer": d.manufacturer,
        "device_type": d.device_type,
        "freq_min_mhz": d.freq_min_mhz,
        "freq_max_mhz": d.freq_max_mhz,
        "package": d.package,
        "package_size": d.package_size,
        "pin_count": d.pin_count,
        "enable_level": d.enable_level,
        "switch_logic": d.switch_logic,
        "bands": d.bands or [],
        "pdf_url": d.pdf_url,
        "created_at": d.created_at.isoformat() if d.created_at else None,
    }


@app.get("/health")
def health():
    return {"status": "ok"}
